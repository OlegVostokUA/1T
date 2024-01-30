import os, sys
import openpyxl
import locale
from datetime import datetime
from PyQt5.QtWidgets import *
from PyQt5 import QtGui


locale.setlocale(locale.LC_ALL, "ru_RU")
date = datetime.today().strftime("%d.%m.%Y")


class Window(QWidget):
    def __init__(self, parent=None):
        self.parent = parent
        super(Window, self).__init__()
        self.make_file = QPushButton('   Почати заповнення')
        self.make_file.setIcon(QtGui.QIcon('img/process.png'))
        self.make_file.clicked.connect(self.scrap_write)

        main_v_box = QVBoxLayout(self)
        main_v_box.addWidget(self.make_file)

    def get_dir_path(self):
        self.files_directory_choice = QFileDialog.getExistingDirectory(self, 'Оберіть папку з вихідними файлами')
        return self.files_directory_choice

    def get_file_path(self):
        self.file_main = QFileDialog.getOpenFileName(self, 'Оберіть базовий файл')
        return self.file_main[0]

    def scrap_write(self):
        f_dir = self.get_dir_path()
        m_file = self.get_file_path()
        files = os.listdir(f_dir)
        val_massive = []
        for file in files:
            read_data = openpyxl.load_workbook(f_dir+'/'+file)
            sheet_1 = read_data.sheetnames[0]
            sheet = read_data[sheet_1]
            columns = [2, 4, 6, 8, 10, 12] # ????

            for coll in columns:
                mark = sheet.cell(column=coll, row=6).value
                if mark == None:
                    break
                val_list = []
                for row in range(6, 119):
                    val = sheet.cell(column=coll, row=row).value
                    if val == None:
                        val = 0
                    if type(val) == float or type(val) == int:
                        val = locale.str(val)
                    val_list.append(val)
                val_massive.append(val_list)

        open_file = openpyxl.load_workbook(m_file)
        sheet_main_1 = open_file.sheetnames[0]
        sheet_main = open_file[sheet_main_1]
        columns_main = [x for x in range(4, (len(val_massive)+2)*2) if x % 2 == 0]

        col_index = 0
        for val_l in val_massive:
            column_numb = columns_main[col_index]
            row = 6
            for val in val_l:
                sheet_main.cell(column=column_numb, row=row).value = val
                row += 1
            col_index += 1
        # variant. But ...
        new_filename = m_file.split('.')
        new_filename.insert(1, date)
        new_filename[2] = '.xlsx'
        new_filename = ' '.join(new_filename)
        open_file.save(new_filename)


class WindowAuthor(QWidget):
    def __init__(self, parent=None):
        self.parent = parent
        super(WindowAuthor, self).__init__()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__() # initialization widgets and properties Parent class "QDialog"
        self.setWindowTitle('Документ Парсер')
        self.setWindowIcon(QtGui.QIcon('img/software.png'))
        self.resize(600, 300) # set size window
        self.main_widget = QTabWidget()
        self.setCentralWidget(self.main_widget)
        self.main_widget.addTab(Window(), "Головне вікно")


if __name__ == '__main__':
    app = QApplication(sys.argv) # create app
    dlgMain = MainWindow() # build object of class "DlgMain" and set here in variable "dlgMain"
    dlgMain.show() # show function
    sys.exit(app.exec_())  # loop app in "sys.exit" func for check logs
